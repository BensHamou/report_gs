from account.decorators import admin_only_required, getRedirectionURL, admin_or_gs_required, can_view_move_required, admin_required, admin_or_validator_required
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.forms import modelformset_factory
from django.core.paginator import Paginator
from django.contrib import messages
from django.db import transaction
from django.urls import reverse
from .filters import *
from .models import *
from account.models import *
from .forms import *
import qrcode
from datetime import date, datetime
from .utils import getMProducts
from django.utils.timezone import now, timedelta
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin
from report.cron import send_stock
from django.templatetags.static import static
from django.utils import timezone


# PACKING

@login_required(login_url='login')
@admin_only_required
def listPackingView(request):
    packings = Packing.objects.all().order_by('-date_modified')
    filteredData = PackingFilter(request.GET, queryset=packings)
    packings = filteredData.qs
    paginator = Paginator(packings, request.GET.get('page_size', 12))
    page_number = request.GET.get('page')
    page = paginator.get_page(page_number)
    
    context = {'page': page, 'filteredData': filteredData}
    return render(request, 'list_packings.html', context)

@login_required(login_url='login')
@admin_only_required
def deletePackingView(request, id):
    packing = get_object_or_404(Packing, id=id)
    try:
        packing.delete()
        url_path = reverse('packings')
        return redirect(getRedirectionURL(request, url_path))
    except Exception as e:
        messages.error(request, f"Erreur lors de la suppression de conditionnement : {e}")
        return redirect(getRedirectionURL(request, reverse('packings')))

@login_required(login_url='login')
@admin_only_required
def createPackingView(request):
    form = PackingForm()
    if request.method == 'POST':
        form = PackingForm(request.POST)
        if form.is_valid():
            form.save(user=request.user)
            url_path = reverse('packings')
            return redirect(getRedirectionURL(request, url_path))
        else:
            messages.error(request, "Veuillez corriger les erreurs ci-dessous.")
    
    context = {'form': form}
    return render(request, 'packing_form.html', context)

@login_required(login_url='login')
@admin_only_required
def editPackingView(request, id):
    packing = get_object_or_404(Packing, id=id)
    form = PackingForm(instance=packing)
    
    if request.method == 'POST':
        form = PackingForm(request.POST, instance=packing)
        if form.is_valid():
            form.save(user=request.user)
            url_path = reverse('packings')
            return redirect(getRedirectionURL(request, url_path))
        else:
            messages.error(request, "Veuillez corriger les erreurs ci-dessous.")
    
    context = {'form': form, 'packing': packing}
    return render(request, 'packing_form.html', context)

# FAMILY

@login_required(login_url='login')
@admin_only_required
def listFamilyView(request):
    families = Family.objects.all().order_by('-date_modified')
    filteredData = FamilyFilter(request.GET, queryset=families)
    families = filteredData.qs
    page_size_param = request.GET.get('page_size')
    page_size = int(page_size_param) if page_size_param else 12
    paginator = Paginator(families, page_size)
    page_number = request.GET.get('page')
    page = paginator.get_page(page_number)
    context = {'page': page, 'filteredData': filteredData}
    return render(request, 'list_families.html', context)

@login_required(login_url='login')
@admin_only_required
def deleteFamilyView(request, id):
    family = get_object_or_404(Family, id=id)
    try:
        family.delete()
        messages.success(request, "Famille supprimée avec succès.")
    except Exception as e:
        messages.error(request, f"Erreur lors de la suppression de la famille : {e}")
    return redirect(getRedirectionURL(request, reverse('families')))

@login_required(login_url='login')
@admin_only_required
def createFamilyView(request):
    form = FamilyForm()
    if request.method == 'POST':
        form = FamilyForm(request.POST, request.FILES)
        if form.is_valid():
            form.save(user=request.user)
            messages.success(request, "Famille créée avec succès.")
            return redirect(getRedirectionURL(request, reverse('families')))
        else:
            messages.error(request, "Veuillez corriger les erreurs ci-dessous.")
    context = {'form': form}
    return render(request, 'family_form.html', context)

@login_required(login_url='login')
@admin_only_required
def editFamilyView(request, id):
    family = get_object_or_404(Family, id=id)
    form = FamilyForm(instance=family)
    if request.method == 'POST':
        form = FamilyForm(request.POST, request.FILES, instance=family)
        if form.is_valid():
            form.save(user=request.user)
            messages.success(request, "Famille mise à jour avec succès.")
            return redirect(getRedirectionURL(request, reverse('families')))
        else:
            messages.error(request, "Veuillez corriger les erreurs ci-dessous.")
    context = {'form': form, 'family': family}
    return render(request, 'family_form.html', context)

# PRODUCT

@login_required(login_url='login')
@admin_only_required
def listProductView(request):
    products = Product.objects.filter(type='Produit Fini').order_by('-date_modified')
    sites = Line.objects.filter(id__in=request.user.lines.all()).values('site_id', 'site__designation').distinct()

    selected_site_id = request.GET.get('site')
    selected_site = None

    if selected_site_id:
        selected_site = Site.objects.get(id=selected_site_id)
    else:
        selected_site = request.user.default_site

    filteredData = ProductFilter(request.GET, queryset=products)
    products = filteredData.qs

    page_size_param = request.GET.get('page_size')
    page_size = int(page_size_param) if page_size_param else 12
    paginator = Paginator(products, page_size)
    page_number = request.GET.get('page')
    page = paginator.get_page(page_number)

    products_with_stock = []
    if selected_site:
        for product in page:
            products_with_stock.append({
                'id': product.id,
                'designation': product.designation,
                'tn_qte': product.tn_qte(selected_site.id),
                'state_stock': product.state_stock(selected_site.id),
                'family': product.family.designation if product.family else '/',
                'type': product.type,
            })
    else:
        products_with_stock = [{'id': product.id, 'designation': product.designation, 'tn_qte': None, 'state_stock': 'Site non sélectionné', 'family': None} for product in page]

    context = {
        'page': page,
        'products_with_stock': products_with_stock,
        'filteredData': filteredData,
        'allowed_sites': sites,
        'selected_site': selected_site,
        'default_site': request.user.default_site,
        'is_pf': True
    }
    return render(request, 'list_products.html', context)

@login_required(login_url='login')
@admin_only_required
def deleteProductView(request, id):
    product = get_object_or_404(Product, id=id)
    try:
        product.delete()
        messages.success(request, "Produit supprimée avec succès.")
    except Exception as e:
        messages.error(request, f"Erreur lors de la suppression du produit : {e}")
    return redirect(getRedirectionURL(request, reverse('products')))

@login_required(login_url='login')
@admin_only_required
def createProductView(request):
    form = ProductForm()
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            form.save(user=request.user)
            messages.success(request, "Produit créé avec succès.")
            return redirect(getRedirectionURL(request, reverse('products')))
        else:
            messages.error(request, "Veuillez corriger les erreurs ci-dessous.")
    
    context = {'form': form, 'is_pf': True}
    return render(request, 'product_form.html', context)

@login_required(login_url='login')
@admin_only_required
def editProductView(request, id):
    product = get_object_or_404(Product, id=id)
    form = ProductForm(instance=product)
    
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            form.save(user=request.user)
            messages.success(request, "Produit mis à jour avec succès.")
            return redirect(getRedirectionURL(request, reverse('products')))
        else:
            messages.error(request, "Veuillez corriger les erreurs ci-dessous.")
    
    context = {'form': form, 'product': product, 'is_pf': True}
    return render(request, 'product_form.html', context)


# MP

@login_required(login_url='login')
@admin_only_required
def listMProductView(request):
    products = Product.objects.filter(type='Matière Première').order_by('-date_modified')
    sites = Line.objects.filter(id__in=request.user.lines.all()).values('site_id', 'site__designation').distinct()

    selected_site_id = request.GET.get('site')
    selected_site = None

    if selected_site_id:
        selected_site = Site.objects.get(id=selected_site_id)
    else:
        selected_site = request.user.default_site

    filteredData = ProductFilter(request.GET, queryset=products)
    products = filteredData.qs

    page_size_param = request.GET.get('page_size')
    page_size = int(page_size_param) if page_size_param else 12
    paginator = Paginator(products, page_size)
    page_number = request.GET.get('page')
    page = paginator.get_page(page_number)

    products_with_stock = []
    if selected_site:
        for product in page:
            products_with_stock.append({
                'id': product.id,
                'designation': product.designation,
                'tn_qte': product.tn_qte(selected_site.id),
                'state_stock': product.state_stock(selected_site.id),
                'last_entry_date': product.last_entry_date(selected_site.id),
                'type': product.type,
            })
    else:
        products_with_stock = [{'id': product.id, 'designation': product.designation, 'tn_qte': None, 'state_stock': 'Site non sélectionné', 
                                'last_entry_date': None } for product in page]

    context = {
        'page': page,
        'products_with_stock': products_with_stock,
        'filteredData': filteredData,
        'allowed_sites': sites,
        'selected_site': selected_site,
        'default_site': request.user.default_site,
        'is_pf': False
    }
    return render(request, 'list_products.html', context)

@login_required(login_url='login')
@admin_only_required
def editMProductView(request, id):
    product = get_object_or_404(Product, id=id)
    form = MProductForm(instance=product)
    
    if request.method == 'POST':
        form = MProductForm(request.POST, instance=product)
        if form.is_valid():
            form.save(user=request.user)
            messages.success(request, "Produit mis à jour avec succès.")
            return redirect(getRedirectionURL(request, reverse('mproducts')))
        else:
            messages.error(request, "Veuillez corriger les erreurs ci-dessous.")
    
    context = {'form': form, 'product': product, 'is_pf': False}
    return render(request, 'product_form.html', context)

@login_required(login_url='login')
@admin_only_required
def syncMProducts(request):
    if request.method == 'POST':
        try:
            for mp in getMProducts():
                code = f'[{mp[1]}] ' if mp[1] else ''
                designation = f'{code}{mp[3]}'
                odoo_id = mp[0]
                unit_dict = {'3': 1, '11': 2, '8': 3, '22': 4, '32': 5, '1': 6}
                product, created = Product.objects.update_or_create(
                    odoo_id=odoo_id,
                    defaults={'designation': designation, 'type': 'Matière Première', 
                              'create_uid': request.user, 'write_uid': request.user, 'packing_id': unit_dict.get(str(mp[2]))})
                if created:
                    product.qte_per_pal = 0
                    product.qte_per_cond = 0
                    product.alert_stock = 0
                    product.alert_stock = 0
                    product.save()
            return JsonResponse({'success': True, 'message': 'Matière Première synchronisés avec succès.'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Erreur pendant la synchronisation: {str(e)}'})
    return JsonResponse({'success': False, 'message': 'Méthode de demande non valide.'})

# MOVES

@login_required(login_url='login')
@admin_or_gs_required
def categories_view(request):
    return render(request, 'categories.html')

# MOVE IN

@login_required(login_url='login')
@can_view_move_required
def list_move(request):
    if request.user.role in ['Admin', 'Validateur']:
        allowed_sites = Line.objects.filter(id__in=request.user.lines.all()).values_list('site_id', flat=True).distinct()
    else:
        allowed_sites = [request.user.default_site.id]

    moves = Move.objects.filter(Q(gestionaire=request.user) | Q(line__in=request.user.lines.all()) |
        Q(line__isnull=True, site__in=allowed_sites)).order_by('-date_modified')


    filteredData = MoveFilter(request.GET, queryset=moves)
    moves = filteredData.qs
    page_size_param = request.GET.get('page_size')
    page_size = int(page_size_param) if page_size_param else 12
    paginator = Paginator(moves, page_size)
    page_number = request.GET.get('page')
    page = paginator.get_page(page_number)
    # today = date.today()
    # # move_out_today = Move.objects.filter(state='Validé', type='Sortie', date=today)
    # # palettes_today = sum(m.palette for m in move_out_today)

    # move_lines = MoveLine.objects.filter(Q(move__state='Validé') & Q(move__type='Sortie') & 
    #                                      (Q(move__line__in=request.user.lines.all()) | Q(move__line__isnull=True, move__site__in=allowed_sites))).select_related('product')
    
    # product_totals = {}
    # for move_line in move_lines:
    #     product_id = move_line.product.id
    #     if product_id not in product_totals:
    #         product_totals[product_id] = {
    #             'designation': move_line.product.designation,
    #             'image': move_line.product.image.url if move_line.product.image else None,
    #             'total_qte': 0,
    #         }
    #     product_totals[product_id]['total_qte'] += move_line.qte
    # top_product = max(product_totals.values(), key=lambda p: p['total_qte'], default=None)
    # active_users_count = User.objects.filter(last_login__gte=now() - timedelta(hours=24), lines__in=request.user.lines.all()).count()


    context = {'page': page, 'filteredData': filteredData }
    return render(request, 'move_list.html', context)

@login_required(login_url='login')
@admin_or_gs_required
def delete_move(request, move_id):
    move = get_object_or_404(Move, id=move_id)
    try:
        move.delete()
        messages.success(request, "Mouvement supprimée avec succès.")
    except Exception as e:
        messages.error(request, f"Erreur lors de la suppression du Mouvement : {e}")
    return redirect(getRedirectionURL(request, reverse('moves')))

# MOVE IN - Produit Fini

@login_required(login_url='login')
@admin_or_gs_required
def families_view(request):
    families = Family.objects.filter(for_mp=False)
    return render(request, 'families.html', {'families': families})

@login_required(login_url='login')
@admin_or_gs_required
def products_view(request, family_id):
    family = get_object_or_404(Family, id=family_id)
    products = Product.objects.filter(family=family, type='Produit Fini')
    return render(request, 'products.html', {'products': products, 'family': family})

@login_required(login_url='login')
@admin_or_gs_required
def create_move_in_view(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    MoveLineDetailFormSet = modelformset_factory(LineDetail, fields=('warehouse', 'emplacement', 'qte'), extra=1)
    
    user = request.user
    user_lines = user.lines.all()
    is_admin = user.role == 'Admin'
    
    default_line = user_lines.first() if user_lines.count() == 1 else None
    show_line_field = user_lines.count() > 1
    if is_admin:
        gestionaires = None if user.lines.count() > 1 else user.lines.first().users.filter(Q(role='Gestionaire') | Q(role='Validateur') | Q(role='Admin'), ~Q(id=1))
    else:
        gestionaires = None

    formset = MoveLineDetailFormSet(queryset=LineDetail.objects.none())

    return render(request, 'form.html', {'product': product, 'formset': formset, 'lines': user_lines if show_line_field else None, 
                                         'gestionaires': gestionaires, 'default_line': default_line, 'is_admin': is_admin, 'show_line_field': show_line_field})

@login_required(login_url='login')
@admin_or_gs_required
def edit_move_in_view(request, move_line_id):
    move_line = get_object_or_404(MoveLine, id=move_line_id)
    move = move_line.move
    line_details = move_line.details.all()
    user = request.user
    user_lines = user.lines.all()
    is_admin = user.role == 'Admin'
    default_line = move_line.move.line or user_lines.first()
    show_line_field = user_lines.count() > 1 and not move.is_transfer and not move.type == 'Sortie'
    gestionaires = default_line.users.filter(Q(role='Gestionaire') | Q(role='Validateur') | Q(role='Admin'), ~Q(id=1))
    default_shifts = default_line.shifts.all()

    context = {
        'move': move_line.move,
        'move_line': move_line,
        'line_details': line_details,
        'product': move_line.product,
        'default_shifts': default_shifts,
        'lines': user_lines if show_line_field else None,
        'gestionaires': gestionaires,
        'default_line': default_line,
        'is_admin': is_admin,
        'show_line_field': show_line_field,
        'warehouses': move_line.move.site.warehouses.all(),
        'emplacements': Emplacement.objects.all()
    }
    return render(request, 'edit_move.html', context)

@login_required(login_url='login')
@admin_or_gs_required
def create_move_pf(request):
    if request.method == 'POST':
        try:
            with transaction.atomic():
                site_id = request.POST.get('site')
                line_id = request.POST.get('line')
                shift_id = request.POST.get('shift')
                gestionaire_id = request.POST.get('gestionaire')
                expiry_date = request.POST.get('expiry_date')
                lot_number = request.POST.get('lot_number')
                production_date = request.POST.get('production_date')
                product = request.POST.get('product')
                observation = request.POST.get('observation', '/')

                production_year = datetime.strptime(production_date, "%Y-%m-%d").year
                if not (site_id and line_id and shift_id and gestionaire_id):
                    return JsonResponse({'success': False, 'message': 'Les champs Ligne, Shift, Date et Gestionaire sont obligatoires.'}, status=200)

                product_obj = Product.objects.get(id=product)
                if product_obj.family.is_expiring and not expiry_date:
                    return JsonResponse({'success': False, 'message': 'Le champ Date d\'expiration est obligatoire.'}, status=200)
                
                palette_total = int(request.POST.get('palette_total', 0))
                existing_move_lines = MoveLine.objects.filter(lot_number=lot_number, move__date__year=production_year, 
                                                              move__line_id=line_id, move__type='Entré', product__type='Produit Fini').exclude(move__state='Annulé')
                existing_total = sum(line.palette for line in existing_move_lines)
                if (existing_total + palette_total) > 180:
                    return JsonResponse({'success': False, 'message': 'Le nombre de palettes pour ce lot est limité à 180.'}, status=200)
                
                move = Move.objects.create(line_id=line_id, site_id=site_id, shift_id=shift_id, gestionaire_id=gestionaire_id, date=production_date,  
                                           state='Brouillon',  type='Entré',  create_uid=request.user, write_uid=request.user)
                move_line = MoveLine.objects.create(lot_number=lot_number, product_id=product, move_id=move.id, create_uid=request.user, 
                                                    expiry_date=expiry_date, write_uid=request.user, observation=observation)
                handleDetails(request, move_line)

                return JsonResponse({'success': True, 'message': 'Entrée créée avec succès.', 'new_record': move_line.move.id}, status=200)

        except Exception as e:
            print(str(e))
            return JsonResponse({'success': False, 'message': f'Erreur lors du traitement de la demande: {str(e)}'}, status=500)

@login_required(login_url='login')
@admin_or_gs_required
def update_move_pf(request, move_line_id):
    if request.method == 'POST':
        try:
            with transaction.atomic():
                site_id = request.POST.get('site', None)
                line_id = request.POST.get('line', None)
                shift_id = request.POST.get('shift', None)
                gestionaire_id = request.POST.get('gestionaire', None)
                lot_number = request.POST.get('lot_number', None)
                production_date = request.POST.get('production_date', False) or None
                expiry_date = request.POST.get('expiry_date', False) or '2099-12-31'
                diff_qte = request.POST.get('diff_qte', 0)
                do_check = int(request.POST.get('do_check', 0))
                observation = request.POST.get('observation', '/')
                move_line = MoveLine.objects.get(id=move_line_id)
                move = Move.objects.get(id=move_line.move.id)

                if move_line.product.family.is_expiring and not request.POST.get('expiry_date', False):
                    return JsonResponse({'success': False, 'message': 'Le champ Date d\'expiration est obligatoire.'}, status=200)
    
                if do_check == 0:
                    production_year = datetime.strptime(production_date, "%Y-%m-%d").year
                    if not (line_id and gestionaire_id):
                        return JsonResponse({'success': False, 'message': 'Les champs Ligne, Date et Gestionaire sont obligatoire.'}, status=200)
                
                    palette_total = int(request.POST.get('palette_total', 0))
                    existing_move_lines = MoveLine.objects.filter(lot_number=lot_number, move__date__year=production_year, 
                                                                move__line_id=line_id, move__type='Entré', product__type='Produit Fini').exclude(move__state='Annulé').exclude(id=move_line_id)
                    existing_total = sum(line.palette for line in existing_move_lines)
                    if (existing_total + palette_total) > 180:
                        return JsonResponse({'success': False, 'message': 'Le nombre de palettes pour ce lot est limité à 180.'}, status=200)
                
                    if shift_id:
                        move.shift_id = shift_id
                    else:
                        return JsonResponse({'success': False, 'message': 'Le champs Shift est obligatoire.'}, status=200)
                    
                if not move.is_transfer:
                    move.site_id = site_id
                    move.line_id = line_id
                    move.gestionaire_id = gestionaire_id
                    move.date = production_date
                    move_line.lot_number = lot_number
                
                handleDetails(request, move_line)
                move.write_uid = request.user
                move.save()
                move_line.write_uid = request.user
                move_line.diff_qte = diff_qte
                move_line.expiry_date = expiry_date
                move_line.observation = observation
                move_line.save()
                return JsonResponse({'success': True, 'message': 'Entrée mise à jour avec succès.'}, status=200)
        except Move.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Entré non trouvée.'}, status=200)
        except MoveLine.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Entré non trouvée.'}, status=200)
        except Exception as e:
            print(str(e))
            return JsonResponse({'success': False, 'message': f'Erreur lors du traitement de la demande: {str(e)}'}, status=500)

# MOVE IN - Matière Première

@login_required(login_url='login')
@admin_or_gs_required
def create_move_in_mp_view(request):
    products = Product.objects.filter(type='Matière Première')
    default_site = request.user.default_site
    default_line = default_site.lines.first()
    MoveLineDetailFormSet = modelformset_factory(LineDetail, fields=('warehouse', 'emplacement', 'qte'), extra=1)
    formset = MoveLineDetailFormSet(queryset=LineDetail.objects.none())
    return render(request, 'form_mp.html', {'products': products, 'formset': formset, 'default_site': default_site.id, 'default_line': default_line.id})

@login_required(login_url='login')
@admin_or_gs_required
def edit_move_in_mp_view(request, move_line_id):
    move_line = get_object_or_404(MoveLine, id=move_line_id)
    products = Product.objects.filter(type='Matière Première')
    line_details = move_line.details.all()
    default_site = move_line.move.site
    default_line = default_site.lines.first()
    return render(request, 'edit_move_mp.html', {'products': products, 'default_site': default_site.id, 'move_line': move_line,
                                                 'default_line': default_line.id, 'default_product': move_line.product, 'warehouses': default_site.warehouses.all(),
                                                 'emplacements': Emplacement.objects.all(), 'line_details': line_details})

@login_required(login_url='login')
@admin_or_gs_required
def create_move_mp(request):
    if request.method == 'POST':
        try:
            with transaction.atomic():
                site_id = request.POST.get('site')
                lot_number = request.POST.get('lot_number')
                product = request.POST.get('product')
                observation = request.POST.get('observation', '/')
                production_date = request.POST.get('production_date', False) or None
                expiry_date = request.POST.get('expiry_date', False) or '2099-12-31'
                if not site_id or not lot_number or not product:
                    return JsonResponse({'success': False, 'message': 'Les champs Site, N° Lot et Produit sont obligatoires.'}, status=200)

                product_obj = Product.objects.get(id=product)

                if product_obj.family:
                    if product_obj.family.is_expiring and not request.POST.get('expiry_date', False):
                        return JsonResponse({'success': False, 'message': 'Le champ Date d\'expiration est obligatoire.'}, status=200)
                    
                cu = request.user
                move = Move.objects.create(site_id=site_id, gestionaire=cu, state='Brouillon', type='Entré', create_uid=cu, write_uid=cu, date=production_date)
                move_line = MoveLine.objects.create(lot_number=lot_number, product_id=product, observation=observation, move_id=move.id, 
                                                    expiry_date=expiry_date, create_uid=cu, write_uid=cu)
                handleDetails(request, move_line)
                return JsonResponse({'success': True, 'message': 'Entrée créée avec succès.', 'new_record': move_line.move.id}, status=200)

        except Exception as e:
            print(str(e))
            return JsonResponse({'success': False, 'message': f'Erreur lors du traitement de la demande: {str(e)}'}, status=500)

@login_required(login_url='login')
@admin_or_gs_required
def update_move_mp(request, move_line_id):
    if request.method == 'POST':
        try:
            with transaction.atomic():
                site_id = request.POST.get('site')
                lot_number = request.POST.get('lot_number')
                product = request.POST.get('product')
                observation = request.POST.get('observation', '/')
                production_date = request.POST.get('production_date', False) or None
                expiry_date = request.POST.get('expiry_date', False) or '2099-12-31'
                diff_qte = request.POST.get('diff_qte', 0)
                if not site_id or not lot_number or not product:
                    return JsonResponse({'success': False, 'message': 'Les champs Site, N° Lot et Produit sont obligatoires.'}, status=200)

                move_line = MoveLine.objects.get(id=move_line_id)
                move = Move.objects.get(id=move_line.move.id)
                cu = request.user


                if move_line.product.family:
                    if move_line.product.family.is_expiring and not request.POST.get('expiry_date', False):
                        return JsonResponse({'success': False, 'message': 'Le champ Date d\'expiration est obligatoire.'}, status=200)

                try:

                    move.site_id = site_id
                    move.gestionaire = cu
                    move.write_uid = cu
                    move.date = production_date
                    move.save()

                    if move.type == 'Entré':
                        move_line.lot_number = lot_number

                    move_line.observation = observation
                    move_line.write_uid = cu
                    move_line.diff_qte = diff_qte
                    move_line.expiry_date = expiry_date
                    move_line.save()
                except Exception as e:
                    print(str(e))
                    return JsonResponse({'success': False, 'message': f'Erreur lors de la mise à jour de l\'entrée: {str(e)}'}, status=500)
                handleDetails(request, move_line)
                            
                return JsonResponse({'success': True, 'message': 'Entrée mise à jour avec succès.'}, status=200)
        except Move.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Entré non trouvée.'}, status=200)
        except MoveLine.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Entré non trouvée.'}, status=200)
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Erreur lors du traitement de la demande: {str(e)}'}, status=500)


# VIEWS 

@login_required(login_url='login')
@can_view_move_required        
def move_detail(request, move_id):
    move = get_object_or_404(
        Move.objects.prefetch_related(
            'move_lines__product__packing',
            'move_lines__details__warehouse',
            'move_lines__details__emplacement'
        ),
        id=move_id
    )
    can_edit, can_cancel, can_confirm, can_validate, can_print = False, False, False, False, move.state == 'Validé' and move.type == 'Entré'
    if request.user.role == 'Admin':
        can_edit = move.type == 'Entré' or move.state == 'Brouillon'
        can_cancel = move.state in ['Brouillon', 'Confirmé']
        can_confirm = move.state == 'Brouillon'
        can_validate = move.state == 'Confirmé'
    elif request.user.role == 'Gestionaire' and move.gestionaire == request.user:
        can_edit = move.state == 'Brouillon'
        can_cancel = move.state == 'Brouillon' and (move.type == 'Sotrie' or not move.is_transfer)
        can_confirm = move.state == 'Brouillon'
    elif request.user.role == 'Validateur' and move.site == request.user.default_site:
        can_edit = move.state == 'Brouillon'
        can_confirm = move.state == 'Brouillon'
        can_validate = move.state == 'Confirmé'
        can_cancel = move.state == 'Brouillon'

    context = {'move': move, 'can_edit': can_edit, 'can_cancel': can_cancel, 'can_confirm': can_confirm, 'can_validate': can_validate, 'can_print': can_print}
    return render(request, 'details_move.html', context)

@login_required(login_url='login')
@admin_or_gs_required
def confirmMove(request, move_id):
    if request.method == 'POST':
        try:
            move = Move.objects.get(id=move_id)
        except Move.DoesNotExist:
            messages.success(request, 'Mouvement introuvable')
            return JsonResponse({'success': False, 'message': 'Mouvement introuvable.'})
        
        if move.state != 'Brouillon':
            return JsonResponse({'success': False, 'message': 'Le mouvement doit être à l\'état Brouillon pour être confirmé.'})
        
        if move.type == 'Sortie':
            for ml in move.move_lines.all():
                if abs(ml.qte - ml.initial_qte) > 0.01:
                    return JsonResponse({'success': False, 'message': f'La quantité scannée ({ml.qte}) pour le produit {ml.product.designation} ne correspond pas à la quantité demandée ({ml.initial_qte}).'})
        
        try:
            move.check_can_confirm()
        except ValueError as e:
            return JsonResponse({'success': False, 'message': str(e)})
            
        success = move.changeState(request.user.id, 'Confirmé')
        if success:
            return JsonResponse({'success': True, 'message': 'Mouvement confirmé avec succès.', 'move_id': move_id})
        else:
            return JsonResponse({'success': False, 'message': 'Erreur lors de la confirmation du mouvement.'})
    return JsonResponse({'success': False, 'message': 'Méthode de requête non valide.'})

@login_required(login_url='login')
@admin_or_gs_required
def validateMove(request, move_id):
    if request.method == 'POST':
        try:
            move = Move.objects.get(id=move_id)
        except Move.DoesNotExist:
            messages.success(request, 'Mouvement introuvable')
            return JsonResponse({'success': False, 'message': 'Mouvement introuvable.'})
        
        if move.state != 'Confirmé':
            return JsonResponse({'success': False, 'message': 'Le mouvement doit être à l\'état Confirmé pour être validé.'})
        
        if move.type == 'Sortie':
            unscanned_exists = DetailCode.objects.filter(
                line_detail__move_line__move=move,
                is_scanned=False
            ).exists()
            if unscanned_exists:
                if request.user.role == 'Admin':
                    if request.POST.get('force_scan') == '1':
                        DetailCode.objects.filter(
                            line_detail__move_line__move=move,
                            is_scanned=False
                        ).update(is_scanned=True)
                    else:
                        return JsonResponse({
                            'success': False,
                            'unscanned': True,
                            'message': 'Certaines palettes n\'ont pas encore été scannées. Voulez-vous forcer la validation ?'
                        })
                else:
                    return JsonResponse({'success': False, 'message': 'Impossible de valider ce mouvement car certaines palettes n\'ont pas encore été scannées.'})

        if request.user.role != 'Admin' and not move.is_transfer and not move.is_isolation:
            if hasDraftMoves(move):
                return JsonResponse({'success': False, 'message': 'Il existe des mouvements en brouillon pour ce site.'})
        
        try:
            move.can_validate()
        except ValueError as e:
            return JsonResponse({'success': False, 'message': str(e)})
        
        success = move.changeState(request.user.id, 'Validé')
        if success:
            try:
                success, message = move.do_after_validation(request.user)

                if move.is_transfer and move.is_isolation and move.type == 'Sortie':
                    mirror_move = Move.objects.get(id=move.mirror.id)
                    mirror_move.check_can_confirm()
                    mirror_move.changeState(request.user.id, 'Confirmé')
                    mirror_move.can_validate()
                    mirror_move.changeState(request.user.id, 'Validé')
                    mirror_move.do_after_validation(request.user)
                    return JsonResponse({'success': True, 'message': "Mouvement validée avec succès, idem pour l'entré dans la zone quarataine.", 'move_id': move_id})
                
                return JsonResponse({'success': True, 'message': message, 'move_id': move_id})
            except ValueError as e:
                return JsonResponse({'success': False, 'message': str(e)})
        else:
            return JsonResponse({'success': False, 'message': 'Erreur lors de la validation du mouvement.'})
    return JsonResponse({'success': False, 'message': 'Méthode de requête non valide.'})

@login_required(login_url='login')
@admin_or_gs_required
def cancelMove(request, move_id):
    if request.method == 'POST':
        try:
            move = Move.objects.get(id=move_id)
        except Move.DoesNotExist:
            messages.success(request, 'Mouvement introuvable')
            return JsonResponse({'success': False, 'message': 'Mouvement introuvable.'})
        
        if move.state not in ['Brouillon', 'Confirmé']:
            return JsonResponse({'success': False, 'message': 'Le mouvement doit être à l\'état Brouillon ou Confirmé pour être annulé.'})

        success = move.changeState(request.user.id, 'Annulé')
        if success:
            if move.is_transfer and move.type == 'Entré':
                try:
                    move.cancel_transfer()
                    move.mirror.changeState(request.user.id, 'Annulé')
                except RuntimeError as e:
                    return JsonResponse({'success': False, 'message': str(e)})
            return JsonResponse({'success': True, 'message': 'Mouvement annulé avec succès.', 'move_id': move_id})
        else:
            return JsonResponse({'success': False, 'message': 'Erreur lors de l\'annulation du mouvement.'})
    return JsonResponse({'success': False, 'message': 'Méthode de requête non valide.'})

class EditMoveBLView(LoginRequiredMixin, View):
    template_name = 'edit_move_bl.html'

    def get(self, request, move_id):
        move = get_object_or_404(Move, id=move_id)
        bls = move.bls.all()
        return render(request, self.template_name, {'move': move, 'bls': bls})

    def post(self, request, move_id):
        move = get_object_or_404(Move, id=move_id)

        for bl in move.bls.all():
            bl_id = str(bl.id)
            delete_flag = request.POST.get(f'bl-delete-{bl_id}')
            if delete_flag == 'true':
                bl.delete()
                continue

            numero = request.POST.get(f'bl-numero-{bl_id}')
            if not numero:
                continue
            is_annexe = request.POST.get(f'bl-is_annexe-{bl_id}') == 'on'
            bl.numero = numero
            bl.is_annexe = is_annexe
            bl.save()

        new_numero = request.POST.get('new-bl-numero')
        new_is_annexe = request.POST.get('new-bl-is_annexe') == 'on'
        if new_numero:
            MoveBL.objects.create(move=move, numero=new_numero, is_annexe=new_is_annexe)
        return redirect(reverse('move_detail', args=[move.id]))
    

@login_required(login_url='login')
@admin_only_required
def sendStockState(request):
    try:
        send_stock(include_qrt=False)
        return JsonResponse({'success': True, 'message': 'E-mail envoyé avec succès'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required(login_url='login')
@admin_only_required
def sendQRTStockState(request):
    try:
        send_stock(include_qrt=True)
        return JsonResponse({'success': True, 'message': 'E-mail envoyé avec succès'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})
    
# FETCH JSON

@login_required(login_url='login')
@admin_or_gs_required
def get_shifts_and_users_for_line(request):
    line_id = request.GET.get('line_id')
    line = get_object_or_404(Line, id=line_id)
    
    shifts = line.shifts.all()
    users = line.users.filter(Q(role='Gestionaire') | Q(role='Validateur') | Q(role='Admin'), ~Q(id=1))
    
    shift_data = [{'id': shift.id, 'name': shift.designation} for shift in shifts]
    user_data = [{'id': user.id, 'name': user.fullname} for user in users]
    
    return JsonResponse({'shifts': shift_data, 'users': user_data, 'site': line.site.id})

@login_required(login_url='login')
@admin_or_gs_required
def get_warehouses_for_line(request):
    line_id = request.GET.get('line_id')
    line = Line.objects.get(id=line_id)
    warehouses = Warehouse.objects.filter(site=line.site)
    warehouse_data = [{'id': warehouse.id, 'name': warehouse.designation} for warehouse in warehouses]

    return JsonResponse({'warehouses': warehouse_data, 'site_id': line.site.id})

@login_required(login_url='login')
@admin_or_gs_required
def get_warehouses_for_site(request):
    site_id = request.GET.get('site_id')
    warehouses = Warehouse.objects.filter(site__id=site_id)
    warehouse_data = [{'id': warehouse.id, 'name': warehouse.designation} for warehouse in warehouses]
    return JsonResponse({'warehouses': warehouse_data})

@login_required(login_url='login')
@admin_or_gs_required
def get_emplacements_for_warehouse(request):
    warehouse_id = request.GET.get('warehouse_id')
    warehouse = Warehouse.objects.get(id=warehouse_id)
    emplacements = Emplacement.objects.filter(warehouse=warehouse)
    emplacement_data = [{'id': emplacement.id, 'name': emplacement.designation} for emplacement in emplacements]

    return JsonResponse({'emplacements': emplacement_data})

@login_required(login_url='login')
@admin_or_gs_required
def generateQRCode(request, detail_id):
    try:
        use_dc = request.GET.get('use_dc') == '1'
        use_dl = request.GET.get('use_dl') == '1'
        
        if use_dc:
            dc = DetailCode.objects.get(id=detail_id)
            qr_data = dc.code
        elif use_dl:
            dl = DisponibilityLine.objects.get(id=detail_id)
            qr_data = dl.code
        else:
            line_detail = LineDetail.objects.get(id=detail_id)
            qr_data = line_detail.code
            
        qr = qrcode.QRCode(box_size=10, border=4)
        qr.add_data(qr_data)
        qr.make(fit=True)
        img = qr.make_image(fill='black', back_color='white')

        response = HttpResponse(content_type="image/png")
        img.save(response, "PNG")
        return response

    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erreur : {e}'})

@login_required(login_url='login')
@admin_or_gs_required
def mark_printed(request, detail_id):
    use_dc = request.GET.get('use_dc') == '1'
    use_dl = request.GET.get('use_dl') == '1'
    if use_dc:
        dc = get_object_or_404(DetailCode, id=detail_id)
        dc.is_printed = True
        dc.save()
        DisponibilityLine.objects.filter(code=dc.code).update(is_printed=True)
    elif use_dl:
        dl = get_object_or_404(DisponibilityLine, id=detail_id)
        dl.is_printed = True
        dl.save()
        DetailCode.objects.filter(code=dl.code).update(is_printed=True)
    else:
        line_detail = get_object_or_404(LineDetail, id=detail_id)
        line_detail.detail_codes.update(is_printed=True)
        codes = line_detail.detail_codes.values_list('code', flat=True)
        DisponibilityLine.objects.filter(code__in=codes).update(is_printed=True)
    return JsonResponse({'success': True})

@login_required(login_url='login')
@admin_only_required
def replace_damaged_palette(request, line_id):
    if request.method == 'POST':
        try:
            line = get_object_or_404(DisponibilityLine, id=line_id)
            new_line = line.replace_damaged()
            return JsonResponse({
                'success': True,
                'message': f"La palette a été remplacée. Nouveau code généré : {new_line.code}",
                'new_code': new_line.code
            })
        except ValueError as e:
            return JsonResponse({'success': False, 'message': str(e)})
        except Exception as e:
            return JsonResponse({'success': False, 'message': f"Erreur lors du remplacement : {e}"})
    return JsonResponse({'success': False, 'message': "Méthode non autorisée."})

def handleDetails(request, move_line):
    n_lot = move_line.n_lot
    if not len(move_line.details.all()) == 0:
        existing_line_details = LineDetail.objects.filter(move_line=move_line)
        existing_ids = [str(detail.id) for detail in existing_line_details]
        to_update_rows = [key.split('_')[2] for key in request.POST.keys() if key.startswith('detail_id_') and request.POST[key]]
        to_add_rows = [key.split('_')[2] for key in request.POST.keys() if key.startswith('detail_id_') and not request.POST[key]]
        to_update_ids = [request.POST.get(f'detail_id_{row_id}') for row_id in to_update_rows]
        to_delete_ids = [detail_id for detail_id in existing_ids if detail_id not in to_update_ids]

        LineDetail.objects.filter(id__in=to_delete_ids).delete()

        for row_id in to_update_rows:
            detail_id = request.POST.get(f'detail_id_{row_id}')
            warehouse_id = request.POST.get(f'warehouse_{row_id}')
            emplacement_id = request.POST.get(f'emplacement_{row_id}')
            palette = request.POST.get(f'palette_{row_id}')
            qte = request.POST.get(f'qte_{row_id}')
            if warehouse_id and emplacement_id and qte and palette:
                LineDetail.objects.filter(id=detail_id).update(warehouse_id=warehouse_id, emplacement_id=emplacement_id, qte=float(qte),
                                                               palette=int(palette), write_uid=request.user, n_lot=n_lot)

        for row_id in to_add_rows:
            warehouse_id = request.POST.get(f'warehouse_{row_id}')
            emplacement_id = request.POST.get(f'emplacement_{row_id}')
            palette = request.POST.get(f'palette_{row_id}')
            qte = request.POST.get(f'qte_{row_id}')

            if warehouse_id and emplacement_id and qte and palette:
                LineDetail.objects.create(move_line=move_line, warehouse_id=warehouse_id, emplacement_id=emplacement_id, n_lot=n_lot, 
                                          qte=float(qte), palette=int(palette), write_uid=request.user, create_uid=move_line.create_uid)
    else:
        row_ids = [key.split('_')[1] for key in request.POST.keys() if key.startswith('warehouse_')]
        for row_id in row_ids:
            warehouse_id = request.POST.get(f'warehouse_{row_id}')
            emplacement_id = request.POST.get(f'emplacement_{row_id}')
            palette = request.POST.get(f'palette_{row_id}')
            qte = request.POST.get(f'qte_{row_id}')

            if warehouse_id and emplacement_id and qte and palette:
                line_detail = LineDetail.objects.create(move_line=move_line, warehouse_id=warehouse_id, emplacement_id=emplacement_id, 
                                                        n_lot=n_lot, qte=float(qte), palette=int(palette), create_uid=request.user, write_uid=request.user)
                
# STOCK

@login_required(login_url='login')
@admin_or_validator_required
def stockDetailView(request, id):
    stock = get_object_or_404(Disponibility, id=id)
    context = {'stock': stock}
    return render(request, 'dispo_detail.html', context)

@login_required(login_url='login')
@admin_or_validator_required
def listStockView(request):
    stocks = Disponibility.objects.all().order_by('-date_modified')
    filteredData = DisponibilityFilter(request.GET, queryset=stocks)
    stocks = filteredData.qs
    
    total_qte = stocks.aggregate(Sum('qte'))['qte__sum'] or 0
    total_palettes = DisponibilityLine.objects.filter(disponibility__in=stocks, status='Valide', qte__gt=0).count()

    paginator = Paginator(stocks, request.GET.get('page_size', 12))
    page_number = request.GET.get('page')
    page = paginator.get_page(page_number)

    context = {'page': page, 'filteredData': filteredData, 'total_qte': total_qte, 'total_palettes': total_palettes}
    return render(request, 'list_dispo.html', context)

@login_required(login_url='login')
@admin_only_required
def deleteStockView(request, id):
    stock = get_object_or_404(Disponibility, id=id)
    try:
        stock.delete()
        url_path = reverse('stocks')
        return redirect(getRedirectionURL(request, url_path))
    except Exception as e:
        messages.error(request, f"Erreur lors de la suppression de stock : {e}")
        return redirect(getRedirectionURL(request, reverse('stocks')))

@login_required(login_url='login')
@admin_only_required
def createStockView(request):
    form = DisponibilityForm()
    if request.method == 'POST':
        form = DisponibilityForm(request.POST)
        if form.is_valid():
            form.save(user=request.user)
            url_path = reverse('stocks')
            return redirect(getRedirectionURL(request, url_path))
        else:
            messages.error(request, "Veuillez corriger les erreurs ci-dessous.")
    
    context = {'form': form}
    return render(request, 'dispo_form.html', context)

@login_required(login_url='login')
@admin_only_required
def editStockView(request, id):
    stock = get_object_or_404(Disponibility, id=id)
    form = EditDisponibilityForm(instance=stock)
    
    if request.method == 'POST':
        form = EditDisponibilityForm(request.POST, instance=stock)
        if form.is_valid():
            form.save(user=request.user)
            
            total_qte = 0
            has_error = False
            for line in stock.lines.all():
                qte_str = request.POST.get(f'line_qte_{line.id}')
                if qte_str is not None:
                    try:
                        new_qte = float(qte_str)
                        if new_qte >= 0:
                            if stock.product.type != 'Matière Première':
                                if new_qte > stock.product.qte_per_pal:
                                    messages.error(request, f"La quantité de la palette {line.sequence} dépasse le maximum ({stock.product.qte_per_pal}).")
                                    has_error = True
                                    continue
                                if stock.product.qte_per_cond > 0 and (new_qte % stock.product.qte_per_cond) != 0:
                                    messages.error(request, f"La quantité de la palette {line.sequence} doit être un multiple de {stock.product.qte_per_cond}.")
                                    has_error = True
                                    continue
                            line.qte = new_qte
                            line.save(update_fields=['qte'])
                        total_qte += new_qte
                    except ValueError:
                        total_qte += line.qte
                else:
                    total_qte += line.qte
            
            if has_error:
                return redirect(request.path)
                
            stock.qte = total_qte
            stock.save(update_fields=['qte'])
            
            url_path = reverse('stock_detail', args=[stock.id])
            return redirect(getRedirectionURL(request, url_path))
        else:
            messages.error(request, "Veuillez corriger les erreurs ci-dessous.")
    
    context = {'form': form, 'stock': stock}
    return render(request, 'dispo_form.html', context)

def extractStockView(request):
    site = request.GET.get('site')
    warehouse = request.GET.get('warehouse')
    emplacement = request.GET.get('emplacement')
    n_lot = request.GET.get('n_lot')
    product = request.GET.get('product')

    queryset = Disponibility.objects.select_related(
        'emplacement__warehouse__site',
        'product'
    ).all().order_by('-emplacement__warehouse__site__designation', 'emplacement__warehouse__designation', 'emplacement__designation', 'product__designation')
    if site:
        queryset = queryset.filter(emplacement__warehouse__site__designation__icontains=site)
    if warehouse:
        queryset = queryset.filter(emplacement__warehouse__designation__icontains=warehouse)
    if emplacement:
        queryset = queryset.filter(emplacement__designation__icontains=emplacement)
    if n_lot:
        queryset = queryset.filter(n_lot__icontains=n_lot)
    if product:
        queryset = queryset.filter(product__designation__icontains=product)
    
    wb = Workbook()
    ws = wb.active
    filename = f"État Stock {timezone.now().strftime('%Y-%m-%d')}"
    ws.title = "Stock"

    header_text = f"Extraction du Stock {timezone.now().strftime('%Y-%m-%d')} - GrupoPuma"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    ws["A1"] = header_text
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=42)
    ws["A1"].fill = PatternFill(start_color="151f31", end_color="151f31", fill_type="solid")

    ws.row_dimensions[2].height = 10

    headers = ["Site", "Magasin", "Zone", "Produit", "N° Lot", "Quantité", "Palette", "Date Production", "Date Expiration"]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}3"] = header
        ws[f"{col_letter}3"].font = Font(bold=True, color="FFFFFF")
        ws[f"{col_letter}3"].fill = PatternFill(start_color="151f31", end_color="151f31", fill_type="solid")
        ws[f"{col_letter}3"].alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[col_letter].auto_size = True

    for row_num, dispo in enumerate(queryset, 4):
        ws[f"A{row_num}"] = dispo.emplacement.warehouse.site.designation
        ws[f"B{row_num}"] = dispo.emplacement.warehouse.designation
        ws[f"C{row_num}"] = dispo.emplacement.designation
        ws[f"D{row_num}"] = dispo.product.designation
        ws[f"E{row_num}"] = dispo.n_lot or '/'
        ws[f"F{row_num}"] = dispo.qte or '1'
        ws[f"G{row_num}"] = dispo.palette or '1'
        ws[f"H{row_num}"] = dispo.production_date.strftime('%Y-%m-%d') if dispo.production_date else "/"
        ws[f"I{row_num}"] = dispo.expiry_date.strftime('%Y-%m-%d') if dispo.expiry_date else "/"

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    ws.auto_filter.ref = f"A3:I{len(queryset) + 2}"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    wb.save(response)
    return response

def hasDraftMoves(move):
    if not move.site.check_for_drafts:
        return False
    
    if not move.move_lines.filter(product__type='Produit Fini').exists():
        return False

    draft_moves = Move.objects.filter(Q(state='Brouillon') | Q(state='Confirmé'), site=move.site, is_transfer=False, is_isolation=False)
    draft_moves_with_finished_products = draft_moves.filter(move_lines__product__type='Produit Fini').distinct()
    return draft_moves_with_finished_products.exists()

@login_required(login_url='login')
@admin_required
def startInventory(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Méthode de requête non valide.'}, status=405)

    try:
        excel_file = export_current_stock()

        create_move_outs_response = create_move_outs(request.user)
        if not create_move_outs_response['success']:
            return JsonResponse(create_move_outs_response, status=500)

        response = HttpResponse(excel_file, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = f"Etat_Stock_Avant_Vidage_{timezone.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erreur lors du démarrage de l\'inventaire: {str(e)}'}, status=500)

def export_current_stock():
    queryset = Disponibility.objects.all().order_by('-emplacement__warehouse__site__designation', 'emplacement__warehouse__designation', 'emplacement__designation',
    'product__designation').select_related('product', 'emplacement__warehouse__site', 'emplacement__warehouse')

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock"

    header_text = f"Extraction du Stock avant vidage - {timezone.now().strftime('%Y-%m-%d %H:%M')} - GrupoPuma"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    ws["A1"] = header_text
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = PatternFill(start_color="151f31", end_color="151f31", fill_type="solid")

    headers = ["Site", "Magasin", "Zone", "Produit", "N° Lot", "Quantité", "Palette", "Date Production", "Date Expiration"]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}3"] = header
        ws[f"{col_letter}3"].font = Font(bold=True, color="FFFFFF")
        ws[f"{col_letter}3"].fill = PatternFill(start_color="151f31", end_color="151f31", fill_type="solid")
        ws[f"{col_letter}3"].alignment = Alignment(horizontal="center", vertical="center")

    for row_num, dispo in enumerate(queryset, 4):
        ws[f"A{row_num}"] = dispo.emplacement.warehouse.site.designation
        ws[f"B{row_num}"] = dispo.emplacement.warehouse.designation
        ws[f"C{row_num}"] = dispo.emplacement.designation
        ws[f"D{row_num}"] = dispo.product.designation
        ws[f"E{row_num}"] = dispo.n_lot or '/'
        ws[f"F{row_num}"] = dispo.qte or 0
        ws[f"G{row_num}"] = dispo.palette or 0
        ws[f"H{row_num}"] = dispo.production_date.strftime('%Y-%m-%d') if dispo.production_date else "/"
        ws[f"I{row_num}"] = dispo.expiry_date.strftime('%Y-%m-%d') if dispo.expiry_date else "/"

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[col_letter].width = adjusted_width

    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    return output.getvalue()

def create_move_outs(user):
    try:
        disponibilities = Disponibility.objects.all().select_related('product', 'emplacement__warehouse').order_by('emplacement__warehouse')

        warehouses_data = {}
        for dispo in disponibilities:
            warehouse_id = dispo.emplacement.warehouse.id
            if warehouse_id not in warehouses_data:
                warehouses_data[warehouse_id] = {'warehouse': dispo.emplacement.warehouse, 'items': []}
            warehouses_data[warehouse_id]['items'].append(dispo)

        for wahrehouse_id, wahrehouse_data in warehouses_data.items():
            site = wahrehouse_data['warehouse'].site
            move = Move.objects.create(
                site=site, 
                gestionaire=user, 
                date=timezone.now().date(),
                type='Sortie',
                state='Brouillon',
                create_uid=user,
                write_uid=user,
                is_inventory=True
            )

            product_lot_groups = {}
            for dispo in wahrehouse_data['items']:
                key = (dispo.product.id, dispo.n_lot or '/')
                if key not in product_lot_groups:
                    product_lot_groups[key] = {
                        'product': dispo.product,
                        'n_lot': dispo.n_lot,
                        'expiry_date': dispo.expiry_date,
                        'dispos': []
                    }
                product_lot_groups[key]['dispos'].append(dispo)

            for (product_id, n_lot), group in product_lot_groups.items():
                move_line = MoveLine.objects.create(
                    move=move,
                    product=group['product'],
                    lot_number=n_lot or '/',
                    expiry_date=group['expiry_date'],
                    create_uid=user,
                    write_uid=user
                )

                for dispo in group['dispos']:
                    LineDetail.objects.create(
                        move_line=move_line,
                        warehouse=dispo.emplacement.warehouse,
                        emplacement=dispo.emplacement,
                        qte=dispo.qte,
                        palette=dispo.palette,
                        expiry_date=group['expiry_date'],
                        n_lot=n_lot or '/',
                        create_uid=user,
                        write_uid=user
                    )
            
            try:
                move.check_can_confirm()
            except ValueError as e:
                return JsonResponse({'success': False, 'message': str(e)})

            move.changeState(user.id, 'Confirmé')
            try:
                move.can_validate()
            except ValueError as e:
                return JsonResponse({'success': False, 'message': str(e)})
                
            success = move.changeState(user.id, 'Validé')
            if success:
                try:
                    success, message = move.do_after_validation(user)
                except ValueError as e:
                    return JsonResponse({'success': False, 'message': str(e)})

        return {'success': True, 'message': f'Créé {len(wahrehouse_data)} mouvements de sortie'}

    except Exception as e:
        return {'success': False, 'message': f'Erreur lors de la création des mouvements de sortie: {str(e)}'}

def cartographieView(request):
    allowed_sites = Line.objects.filter(id__in=request.user.lines.all()).values_list('site_id', flat=True).distinct()
    sites = Site.objects.prefetch_related(
        'warehouses',
        'warehouses__emplacements',
        'warehouses__emplacements__disponibilities',
        'warehouses__emplacements__disponibilities__product',
        'warehouses__emplacements__disponibilities__product__packing'
    ).filter(
        id__in=allowed_sites
    ).order_by('designation')
    return render(request, 'cartographie.html', {
        'sites': sites,
        'empty_image': static('img/empty_cartography.png')
        })

@login_required(login_url='login')
@admin_or_gs_required
def edit_move_out_line_view(request, move_line_id):
    import json
    from django.db import models
    move_line = get_object_or_404(MoveLine, id=move_line_id)
    if request.method == 'POST':
        try:
            with transaction.atomic():
                initial_qte = request.POST.get('initial_qte')
                observation = request.POST.get('observation', '/')
                deleted_codes_str = request.POST.get('deleted_codes', '[]')
                modified_quantities_str = request.POST.get('modified_quantities', '{}')
                
                if initial_qte:
                    initial_qte_float = float(initial_qte)
                    if initial_qte_float <= 0:
                        return JsonResponse({'success': False, 'message': 'La quantité doit être supérieure à 0.'}, status=200)
                    move_line.initial_qte = initial_qte_float
                
                move_line.observation = observation
                move_line.save()
                
                # Parse JSON
                try:
                    deleted_codes = json.loads(deleted_codes_str)
                    modified_quantities = json.loads(modified_quantities_str)
                except json.JSONDecodeError:
                    return JsonResponse({'success': False, 'message': 'Données de palettes invalides.'}, status=400)
                
                # Delete removed codes
                if deleted_codes:
                    dcs = DetailCode.objects.filter(id__in=deleted_codes, line_detail__move_line=move_line)
                    line_details = set(dc.line_detail for dc in dcs)
                    dcs.delete()
                    
                    for ld in line_details:
                        total = ld.detail_codes.aggregate(total=models.Sum('qte'))['total'] or 0
                        pal = ld.detail_codes.count()
                        if pal == 0:
                            ld.delete()
                        else:
                            ld.qte = total
                            ld.palette = pal
                            ld.save()
                
                # Process quantity updates
                for dc_id_str, new_qte in modified_quantities.items():
                    dc = DetailCode.objects.filter(id=int(dc_id_str), line_detail__move_line=move_line).first()
                    if dc:
                        new_qte_float = float(new_qte)
                        if new_qte_float > 0 and dc.qte != new_qte_float:
                            product = move_line.product
                            if product.qte_per_pal and new_qte_float > product.qte_per_pal:
                                return JsonResponse({'success': False, 'message': f"La quantité ({new_qte_float}) dépasse la limite de {product.qte_per_pal} par palette."}, status=200)
                            
                            if product.qte_per_cond:
                                remainder = new_qte_float % product.qte_per_cond
                                if remainder > 1e-5 and (product.qte_per_cond - remainder) > 1e-5:
                                    return JsonResponse({'success': False, 'message': f"La quantité ({new_qte_float}) n'est pas un multiple de {product.qte_per_cond}."}, status=200)

                            dc.qte = new_qte_float
                            dc.save()
                            
                            # Update parent LineDetail
                            ld = dc.line_detail
                            total = ld.detail_codes.aggregate(total=models.Sum('qte'))['total'] or 0
                            ld.qte = total
                            ld.save()
                
                # MoveLine's qte is a property that calculates itself automatically,
                # so we only need to save the initial_qte and observation changes.
                move_line.save()
                
                return JsonResponse({'success': True, 'message': 'Ligne mise à jour avec succès.', 'move_id': move_line.move.id}, status=200)
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Erreur lors du traitement: {str(e)}'}, status=500)
            
    return render(request, 'move/edit_move_out_line.html', {'move_line': move_line})

@login_required(login_url='login')
@admin_or_gs_required
def get_available_stock(request):
    product_id = request.GET.get('product_id')
    site_id = request.GET.get('site_id')
    move_type = request.GET.get('move_type', 'normal')
    if not product_id or not site_id:
        return JsonResponse({'success': False, 'message': 'Paramètres manquants.'}, status=400)
    
    dispos = Disponibility.objects.filter(
        product_id=product_id,
        emplacement__warehouse__site_id=site_id,
        qte__gt=0
    ).select_related('emplacement', 'emplacement__warehouse')
    
    if move_type == 'normal':
        dispos = dispos.filter(emplacement__quarantine=False, emplacement__temp=False).order_by('expiry_date', 'n_lot')
    elif move_type == 'isolation':
        dispos = dispos.filter(emplacement__quarantine=False)
    elif move_type == 'consumption':
        dispos = dispos.filter(emplacement__quarantine=True)
    elif move_type == 'transfer':
        dispos = dispos.filter(emplacement__quarantine=False, emplacement__temp=False).order_by('expiry_date', 'n_lot')

    stock_list = []
    for d in dispos:
        palettes = []
        for line in d.sorted_lines:
            if line.status == 'Valide':
                palettes.append({
                    'id': line.id,
                    'code': line.code,
                    'qte': line.qte,
                    'sequence': line.sequence
                })
        stock_list.append({
            'emplacement_id': d.emplacement.id,
            'warehouse_id': d.emplacement.warehouse.id,
            'emplacement_name': f"{d.emplacement.warehouse.designation} - {d.emplacement.designation}",
            'n_lot': d.n_lot,
            'qte': d.qte,
            'palette': d.palette,
            'expiry_date': d.expiry_date.strftime('%Y-%m-%d') if d.expiry_date else None,
            'palettes': palettes
        })
    return JsonResponse({'success': True, 'stock': stock_list})

@login_required(login_url='login')
@admin_or_gs_required
def create_move_out_view(request):
    if request.method == 'POST':
        try:
            import json
            try:
                data = json.loads(request.body)
            except json.JSONDecodeError:
                data = request.POST

            with transaction.atomic():
                site_id = data.get('site')
                line_id = data.get('line') or None
                shift_id = data.get('shift') or None
                gestionaire_id = data.get('gestionaire')
                date_str = data.get('date')
                move_type = data.get('move_type', 'normal')
                observation = data.get('observation', '/')
                bl_number = data.get('bl_number')
                btr_number = data.get('btr_number')
                category = data.get('category', 'PF')
                
                is_mp = (category == 'MP')
                is_transfer = move_type in ['transfer', 'isolation']
                is_isolation = move_type in ['isolation', 'consumption']
                transfer_to_id = data.get('transfer_to') if move_type == 'transfer' else None

                if not (site_id and date_str and gestionaire_id):
                    return JsonResponse({'success': False, 'message': 'Les champs Site, Date et Gestionaire sont obligatoires.'}, status=200)

                if move_type == 'normal' and not bl_number and not is_mp:
                    return JsonResponse({'success': False, 'message': 'Le champ N° BL est obligatoire pour les sorties normales.'}, status=200)
                if move_type == 'transfer' and not btr_number:
                    return JsonResponse({'success': False, 'message': 'Le champ N° BTR est obligatoire pour les transferts.'}, status=200)

                products_list = data.get('products', [])
                if not products_list:
                    return JsonResponse({'success': False, 'message': 'Veuillez ajouter au moins un produit à sortir.'}, status=200)

                move = Move.objects.create(
                    site_id=site_id,
                    line_id=line_id,
                    shift_id=shift_id,
                    gestionaire_id=gestionaire_id,
                    date=date_str,
                    type='Sortie',
                    is_transfer=is_transfer,
                    is_isolation=is_isolation,
                    is_mp=is_mp,
                    transfer_to_id=transfer_to_id,
                    state='Brouillon',
                    create_uid=request.user,
                    write_uid=request.user
                )

                if move_type == 'normal':
                    if is_mp:
                        # Auto generate sequence for MP
                        year = int(date_str.split('-')[0]) if date_str else move.date.year
                        seq_obj, created = MoveMPSequence.objects.select_for_update().get_or_create(
                            site_id=site_id, year=year
                        )
                        seq_obj.sequence += 1
                        seq_obj.save()
                        MoveBL.objects.create(move=move, numero=seq_obj.sequence, is_annexe=False)
                    else:
                        try:
                            MoveBL.objects.create(move=move, numero=int(bl_number), is_annexe=False)
                        except (ValueError, TypeError):
                            return JsonResponse({'success': False, 'message': 'Le N° BL doit être un nombre entier.'}, status=200)
                elif move_type == 'transfer' and btr_number:
                    try:
                        MoveBL.objects.create(move=move, numero=int(btr_number), is_annexe=False)
                    except (ValueError, TypeError):
                        return JsonResponse({'success': False, 'message': 'Le N° BTR doit être un nombre entier.'}, status=200)

                for prod_item in products_list:
                    product_id = prod_item.get('product_id')
                    desired_qte = float(prod_item.get('qte', 0))
                    
                    if not product_id or desired_qte <= 0:
                        continue

                    MoveLine.objects.create(
                        move=move,
                        product_id=product_id,
                        lot_number='/',
                        initial_qte=desired_qte,
                        create_uid=request.user,
                        write_uid=request.user,
                        observation=observation
                    )

                if not move.move_lines.exists():
                    move.delete()
                    return JsonResponse({'success': False, 'message': 'Veuillez ajouter au moins un produit valide à sortir.'}, status=200)

                return JsonResponse({'success': True, 'message': 'Mouvement de sortie créé avec succès (Brouillon).', 'new_record': move.id}, status=200)

        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Erreur lors de la création de la sortie : {str(e)}'}, status=500)

    if request.GET.get('fetch_products') == '1':
        mp_products = Product.objects.filter(type='Matière Première').select_related('packing').order_by('designation')
        pf_products = Product.objects.filter(type='Produit Fini').select_related('family', 'packing').order_by('designation')
        
        pf_data = [{
            'id': p.id,
            'family_id': p.family_id,
            'designation': p.designation,
            'qte_per_pal': p.qte_per_pal or 1.0,
            'image_url': p.image.url if p.image else '',
            'qte_per_cond': p.qte_per_cond or 0.0,
            'unit': p.packing.unit if p.packing else 'Kg'
        } for p in pf_products]
        
        mp_data = [{
            'id': p.id,
            'designation': p.designation,
            'qte_per_pal': p.qte_per_pal or 1.0,
            'qte_per_cond': p.qte_per_cond or 0.0,
            'unit': p.packing.unit if p.packing else 'Kg'
        } for p in mp_products]

        return JsonResponse({'pf_products': pf_data, 'mp_products': mp_data})

    user = request.user
    if user.role == 'Admin':
        sites = Site.objects.all()
    elif user.default_site:
        sites = Site.objects.filter(id=user.default_site.id)
    else:
        sites = Site.objects.none()

    families = Family.objects.filter(for_mp=False).order_by('sequence', 'designation')
    lines = user.lines.all()
    gestionaires = User.objects.filter(Q(role='Gestionaire') | Q(role='Admin') | Q(role='Validateur')).exclude(is_superuser=True)
    
    context = {
        'sites': sites,
        'families': families,
        'mp_products': [],
        'pf_products': [],
        'lines': lines,
        'gestionaires': gestionaires,
        'default_site': user.default_site,
        'is_admin': user.role == 'Admin'
    }
    return render(request, 'move/pf/form_sortie.html', context)


@login_required(login_url='login')
@admin_only_required
def extourneMove(request, move_id):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Méthode non autorisée.'})
        
    try:
        move = Move.objects.get(id=move_id)
    except Move.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Mouvement introuvable.'})
        
    if move.state != 'Validé':
        return JsonResponse({'success': False, 'message': 'Seul un mouvement Validé peut être extourné.'})
        
    if move.is_extourne:
        return JsonResponse({'success': False, 'message': 'Ce mouvement a déjà été extourné.'})
        
    if move.is_transfer or move.is_isolation or move.is_inventory:
        return JsonResponse({'success': False, 'message': 'Seuls les rapports d\'Entré ou de Sortie simples peuvent être extournés.'})
        
    try:
        with transaction.atomic():
            new_type = 'Sortie' if move.type == 'Entré' else 'Entré'
            
            extourne = Move.objects.create(
                site=move.site,
                line=move.line,
                shift=move.shift,
                gestionaire=request.user,
                is_mp=move.is_mp,
                type=new_type,
                state='Brouillon',
                extourned_by=move
            )
            
            for ml in move.move_lines.all():
                new_ml = MoveLine.objects.create(
                    move=extourne,
                    product=ml.product,
                    lot_number=ml.lot_number,
                    observation=ml.observation,
                    initial_qte=ml.qte,
                    expiry_date=ml.expiry_date,
                    create_uid=request.user,
                    write_uid=request.user
                )
                
                for ld in ml.details.all():
                    from report.models import Disponibility
                    dispo = Disponibility.objects.filter(product=ml.product, emplacement=ld.emplacement, n_lot=ld.n_lot).first()
                    exp_date = ld.expiry_date or ml.expiry_date or (dispo.expiry_date if dispo else None)
                    
                    new_ld = LineDetail.objects.create(
                        move_line=new_ml,
                        warehouse=ld.warehouse,
                        emplacement=ld.emplacement,
                        n_lot=ld.n_lot,
                        qte=ld.qte,
                        palette=ld.palette,
                        expiry_date=exp_date,
                        create_uid=request.user,
                        write_uid=request.user
                    )
                    
                    if new_type == 'Sortie':
                        for dc in ld.detail_codes.all():
                            DetailCode.objects.create(
                                line_detail=new_ld,
                                code=dc.code,
                                qte=dc.qte,
                                palette=dc.palette,
                                is_scanned=True
                            )
            
            extourne.check_can_confirm()
            extourne.changeState(request.user.id, 'Confirmé')
            
            extourne.can_validate()
            extourne.changeState(request.user.id, 'Validé')
            success, msg = extourne.do_after_validation(request.user)
            if not success:
                raise ValueError(msg)
                
            move.is_extourne = True
            move.save(update_fields=['is_extourne'])
            
            return JsonResponse({'success': True, 'message': f'Le mouvement a été extourné avec succès. Nouveau mouvement ID: {extourne.id}', 'new_move_id': extourne.id})
            
    except ValueError as e:
        return JsonResponse({'success': False, 'message': str(e)})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Erreur inattendue: {str(e)}'})
